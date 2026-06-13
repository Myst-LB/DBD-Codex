#!/usr/bin/env python3
"""
build_data.py  —  Parse xlsx + txt and emit data.js for the DBD Codex site.
Run from repo root:  python scripts/build_data.py
"""
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "dbd_perks_updated_F - Copy.xlsx"
TXT  = ROOT / "dbd_perks_updated_F.txt"
OUT  = ROOT / "data.js"

# ── helpers ────────────────────────────────────────────────────────────────────
def clean(s):
    if s is None:
        return ""
    s = str(s).strip()
    # openpyxl normalises most things; clean up residual mojibake / display chars
    s = s.replace("舒", "—").replace("ÿþ", "").replace("&&&&&", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def cell(row, idx):
    """Safe cell value from an openpyxl row (list)."""
    try:
        return clean(row[idx].value if hasattr(row[idx], "value") else row[idx])
    except IndexError:
        return ""

# ── 1. Read character map from txt (UTF-16, every perk has character filled) ───
char_map = {}   # perk_name_lower → character string
try:
    txt_raw = TXT.read_text(encoding="utf-16")
except Exception as e:
    sys.exit(f"Cannot read txt: {e}")

for line in txt_raw.splitlines():
    parts = line.split("\t")
    # normalize whitespace in each part (collapse runs of spaces, preserve word boundaries)
    parts = [" ".join(p.split()) for p in parts]
    # remove empty + &&&&& separators
    parts = [p for p in parts if p and p != "&&&&&"]
    # data rows start with a digit id
    if not parts or not parts[0].isdigit():
        continue
    if len(parts) < 4:
        continue
    # txt columns: [0]id [1]tier [2]★★★★★ [3]name [4]character [5]category [6]desc [7]synergy [8]origin
    pid   = parts[0]
    name  = parts[3] if len(parts) > 3 else ""
    char  = parts[4] if len(parts) > 4 else ""
    char_map[name.lower()] = char
    char_map[pid] = char   # also index by id as fallback

print(f"[txt] character map built: {len(char_map)} entries")

# ── 2. Load xlsx ────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 3. Parse Full Tier List (sheet1) ────────────────────────────────────────────
ws_perks = wb["Full Tier List"]
perks = []
for row in ws_perks.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    # skip header, banner rows, and empty rows — data rows start with a digit id
    if not vals[0] or not vals[0].isdigit():
        continue
    pid      = int(vals[0])
    tier     = vals[1]
    # col 2 is star rating — skip
    name     = vals[3]
    char_raw = vals[4]   # mostly empty in xlsx
    category = vals[5]
    desc     = vals[6]
    synergy  = vals[7]

    # backfill character from txt
    character = char_raw or char_map.get(name.lower()) or char_map.get(str(pid), "")

    perks.append({
        "id":          pid,
        "tier":        tier,
        "name":        name,
        "character":   character,
        "category":    category,
        "description": desc,
        "synergy":     synergy,
    })

print(f"[xlsx] perks parsed: {len(perks)}")
no_char = [p["name"] for p in perks if not p["character"]]
if no_char:
    print(f"  WARNING: {len(no_char)} perks missing character: {no_char[:5]}…")

# tier sanity
from collections import Counter
tier_counts = Counter(p["tier"] for p in perks)
print(f"  tier counts: {dict(tier_counts)}")
assert len(perks) == 169, f"Expected 169 perks, got {len(perks)}"  # pre-deletion count

# ── 4. Parse Killer Tier List (sheet5) ──────────────────────────────────────────
ws_killers = wb["Killer Tier List"]
killers = []
for row in ws_killers.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    if not vals[0] or not vals[0].isdigit():
        continue
    killers.append({
        "rank":       int(vals[0]),
        "tier":       vals[1],
        "name":       vals[2],
        "power":      vals[3],
        "strengths":  vals[4],
        "weaknesses": vals[5],
        "status":     vals[6],
        "price":      vals[7],
        "perks":      [v for v in vals[8:11] if v],
    })

print(f"[xlsx] killers parsed: {len(killers)}")
assert len(killers) in (41, 42), f"Expected 41-42 killers, got {len(killers)}"

# ── 5. Parse Survivor Tier List (sheet6) ────────────────────────────────────────
ws_survivors = wb["Survivor Tier List"]
survivors = []
for row in ws_survivors.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    if not vals[0] or not vals[0].isdigit():
        continue
    survivors.append({
        "rank":      int(vals[0]),
        "name":      vals[1],
        "modelSize": vals[2],
        "noise":     vals[3],
        "status":    vals[4],
        "price":     vals[5],
        "notes":     vals[6],
        "perks":     [v for v in vals[7:10] if v],
    })

print(f"[xlsx] survivors parsed: {len(survivors)}")
assert len(survivors) == 52, f"Expected 52 survivors, got {len(survivors)}"

# ── 6. Parse Combo Builds (sheet2) ──────────────────────────────────────────────
ws_builds = wb["Combo Builds"]
builds = []
for row in ws_builds.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    if not vals[0] or vals[0] == "Build Name":
        continue
    builds.append({
        "name":     vals[0],
        "perks":    vals[1],
        "strategy": vals[2],
    })

print(f"[xlsx] builds parsed: {len(builds)}")
assert len(builds) == 15, f"Expected 15 builds, got {len(builds)}"

# Fix perk names misspelled in the build text so they resolve to real perks
BUILD_PERK_FIXES = {
    "Flash Grenade": "Flashbang",   # the Flashbang perk (crafts a flash grenade)
}
for b in builds:
    for wrong, right in BUILD_PERK_FIXES.items():
        if b["perks"] and wrong in b["perks"]:
            b["perks"] = b["perks"].replace(wrong, right)

# Same fixes in perk "synergy" lists, which render as clickable perk links
for p in perks:
    syn = p.get("synergy")
    if syn:
        for wrong, right in BUILD_PERK_FIXES.items():
            syn = syn.replace(wrong, right)
        p["synergy"] = syn

# ── 7. Parse Summary / Meta (sheet4) ────────────────────────────────────────────
ws_summary = wb["Summary"]
meta = {"version": "9.5.0", "source": "", "tiers": []}
tier_table_started = False
for row in ws_summary.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    if not any(vals):
        continue
    if vals[0].startswith("Source:"):
        meta["source"] = vals[0].replace("Source:", "").strip()
    if vals[0] in ("Tier",):
        tier_table_started = True
        continue
    if tier_table_started and vals[0] and vals[0][0] == "★":
        meta["tiers"].append({
            "tier":  vals[0],
            "count": vals[1],
            "desc":  vals[2],
        })

# ── 7.5. Apply corrections (fix speech-recognition errors in source transcripts) ────
# Survivor tier list: correct names by rank
SURVIVOR_RANK_CORRECTIONS = {
    45: "Yoichi Asakawa",            # was "Yui Kimura" (speech-rec error)
    29: "Taurie Cain",               # was "Tapp Cain"
    22: "Aestri Yazar & Baermar Uraz",  # was "Ace Trapper"
    2:  "Kwon Tae-young",            # was "Quentin Smith" (different character added in 9.5.0)
    27: "Orela Rose",                # was "Aurélie Dubois" (Steady Pulse Chapter paramedic)
}
for s in survivors:
    if s["rank"] in SURVIVOR_RANK_CORRECTIONS:
        s["name"] = SURVIVOR_RANK_CORRECTIONS[s["rank"]]

# Perk name corrections (bad names in xlsx → correct DBD perk names)
PERK_NAME_MAP = {
    "Stillside":             "Still Sight",
    "Mirror Image Illusion": "Mirrored Illusion",
    "Thrill of the Hunt":    "Clean Break",      # Taurie Cain's 3rd perk (was wrong name)
    "Stamina":               "Rapid Response",   # Orela Rose's 3rd perk (was misheard)
}
for p in perks:
    if p["name"] in PERK_NAME_MAP:
        p["name"] = PERK_NAME_MAP[p["name"]]

# Perk-to-character corrections (some perks ended up under the wrong character)
PERK_CHAR_MAP = {
    # Yoichi Asakawa's perks (were attributed to "Yui Kimura")
    "Boon: Dark Theory":             "Yoichi Asakawa",
    "Empathic Connection":           "Yoichi Asakawa",
    "Parental Guidance":             "Yoichi Asakawa",
    # Taurie Cain's perks (were attributed to "Tapp Cain")
    "Invocation: Treacherous Crows": "Taurie Cain",
    "Shoulder the Burden":           "Taurie Cain",
    "Clean Break":                   "Taurie Cain",
    # Aestri Yazar & Baermar Uraz's perks (were attributed to "Ace Trapper")
    "Bardic Inspiration":            "Aestri Yazar & Baermar Uraz",
    "Still Sight":                   "Aestri Yazar & Baermar Uraz",
    "Mirrored Illusion":             "Aestri Yazar & Baermar Uraz",
    # Kwon Tae-young's perks (were attributed to "Quentin Smith")
    "Five Moves Ahead":              "Kwon Tae-young",
    "Flow State":                    "Kwon Tae-young",
    "A Place for Us":                "Kwon Tae-young",
    # Orela Rose's perks (were attributed to "Aurélie Dubois")
    "Do No Harm":                    "Orela Rose",
    "Duty of Care":                  "Orela Rose",
    "Rapid Response":                "Orela Rose",
    # Attribution fixes
    "Plunderer's Instinct":          "Base game (all)",
    "Diversion":                     "Adam Francis",
    "Lucky Star":                    "Ellen Ripley",
    "Lightfooted":                   "Ellen Ripley",
    "Wake Up":                       "Quentin Smith",
    "Come and Get Me":               "Rick Grimes",
}
for p in perks:
    if p["name"] in PERK_CHAR_MAP:
        p["character"] = PERK_CHAR_MAP[p["name"]]

# Delete fake / duplicate perks that don't exist in the real game
PERKS_TO_DELETE = {"Poise", "Suitwise", "Situation Awareness"}
perks = [p for p in perks if p["name"] not in PERKS_TO_DELETE]
print(f"[corrections] deleted fake perks: {PERKS_TO_DELETE}  (remaining: {len(perks)})")

# Perk description corrections (where the description was also wrong/mismatched)
PERK_DESC_MAP = {
    "Clean Break":    "Activate while being healed: become Broken for 60 s, then auto-heal to full health. "
                      "Ends the healing action immediately. Strong anti-tracking tool in the right build.",
    "Still Sight":    "Stand still for 2 s → see auras of Generators, Chests, and Killer within 24 m. "
                      "Also activates while being healed. Decent info perk; outclassed by Extra Sensory Perception.",
    "Mirrored Illusion": "Stand still briefly near a Generator → spawn a visible copy of yourself on it "
                         "(visible to the Killer). Rarely fools killers for more than a second or two.",
}
for p in perks:
    if p["name"] in PERK_DESC_MAP:
        p["description"] = PERK_DESC_MAP[p["name"]]

print("[corrections] survivor names fixed:", list(SURVIVOR_RANK_CORRECTIONS.values()))
print("[corrections] perk names fixed:", list(PERK_NAME_MAP.keys()))
print("[corrections] perk characters fixed:", len(PERK_CHAR_MAP), "perks")

# ── 7.6. Inject perks missing from source data (added after 9.5.0 recording) ──
INJECTED_PERKS = [
    {
        "id":          170,
        "tier":        "Weak/Niche",
        "name":        "Teamwork: Toughen Up",
        "character":   "Rick Grimes",
        "category":    "Stealth/Team",
        "description": "While injured, whenever another Survivor within 24 m blinds or Pallet-stuns the Killer, "
                       "for 30 s: you make no Grunts of Pain, leave no Pools of Blood, and leave no Scratch Marks. "
                       "Very situational — requires a teammate to land a stun/blind while you are already injured.",
        "synergy":     "",
    },
    {
        "id":          171,
        "tier":        "Decent",
        "name":        "Teamwork: Soft-Spoken",
        "character":   "Eleven",
        "category":    "Generator/Team",
        "description": "For each other Survivor repairing a Generator with you, the range of Generator repair noises "
                       "is 25% smaller. While repairing with at least 1 other Survivor, you repair 5% faster. "
                       "Minor but consistent passive bonus on co-op gens.",
        "synergy":     "",
    },
]
perks.extend(INJECTED_PERKS)
print(f"[injected] {len(INJECTED_PERKS)} missing perks added: {[p['name'] for p in INJECTED_PERKS]}")

# ── 7.7. Killer Perks (injected from killer_perks_data.py) ─────────────────
sys.path.insert(0, str(Path(__file__).parent))
from killer_perks_data import get_killer_perks
killer_perks = get_killer_perks()
print(f"[killer_perks] loaded {len(killer_perks)} killer perks")

# ── 7.8. Killer Perk Value Index ────────────────────────────────────────────
from collections import defaultdict as _dd2

TIER_SCORE_K = {"Excellent": 5, "Very Good": 4, "Decent": 3, "Weak/Niche": 1, "Terrible": 0}

killer_status_map = {k["name"].lower(): k["status"] for k in killers}
killer_price_map  = {k["name"].lower(): k["price"]  for k in killers}

kp_by_char = _dd2(list)
for p in killer_perks:
    kp_by_char[p["character"] or "Unknown"].append(p)

killer_char_index = []
for char_name, char_kperks in kp_by_char.items():
    tier_score  = sum(TIER_SCORE_K.get(p["tier"], 0) for p in char_kperks)
    total_score = tier_score

    is_general = char_name == "General (all killers)"
    status = "Free" if is_general else killer_status_map.get(char_name.lower(), "Paid")
    price  = "" if is_general else killer_price_map.get(char_name.lower(), "")

    best = max(char_kperks, key=lambda p: TIER_SCORE_K.get(p["tier"], 0))

    killer_char_index.append({
        "name":       char_name,
        "status":     status,
        "price":      price,
        "perkCount":  len(char_kperks),
        "tierScore":  tier_score,
        "totalScore": total_score,
        "bestPerk":   best["name"],
        "bestTier":   best["tier"],
        "perks": sorted([
            {
                "id":        p["id"],
                "name":      p["name"],
                "tier":      p["tier"],
                "tierScore": TIER_SCORE_K.get(p["tier"], 0),
                "category":  p["category"],
            }
            for p in char_kperks
        ], key=lambda p: -p["tierScore"]),
    })

killer_char_index.sort(key=lambda c: -c["totalScore"])
krank = 1
for c in killer_char_index:
    if c["name"] == "General (all killers)":
        c["rank"] = 0
    else:
        c["rank"] = krank
        krank += 1

print(f"[analysis] killer char value index: {len(killer_char_index)} entries")
print(f"  top 5: " + ", ".join(f"{c['name']} ({c['totalScore']})" for c in killer_char_index[:5]))

# ── 8. Compute Perk Value Index ─────────────────────────────────────────────────
from collections import defaultdict

TIER_SCORE = {"Excellent": 5, "Very Good": 4, "Decent": 3, "Weak/Niche": 1, "Terrible": 0}

# Count how many times each perk name is cited in ANY other perk's synergy field
synergy_mentions = Counter()  # perk_name_lower → count
for p in perks:
    if not p["synergy"]:
        continue
    for part in p["synergy"].split(","):
        raw = re.match(r"^([^(]+)", part.strip())
        if raw:
            synergy_mentions[raw.group(1).strip().lower()] += 1

# Per-perk synergy count (cited-by count in the perk data itself)
for p in perks:
    p["synergyCount"] = synergy_mentions.get(p["name"].lower(), 0)

# Build survivor status lookup (name → status)
survivor_status = {s["name"].lower(): s["status"] for s in survivors}
survivor_price  = {s["name"].lower(): s["price"]  for s in survivors}

# Group perks by character
by_char = defaultdict(list)
for p in perks:
    by_char[p["character"] or "Unknown"].append(p)

# Compute score per character
char_index = []
for char_name, char_perks in by_char.items():
    tier_score    = sum(TIER_SCORE.get(p["tier"], 0) for p in char_perks)
    synergy_score = sum(p["synergyCount"] for p in char_perks)
    total_score   = tier_score + synergy_score

    # Status from survivors sheet; Base game characters are always free
    is_base = "base game" in char_name.lower()
    status  = "Free" if is_base else survivor_status.get(char_name.lower(), "Paid")
    price   = "" if is_base else survivor_price.get(char_name.lower(), "")

    # Best perk by tier then synergy
    best = max(char_perks, key=lambda p: (TIER_SCORE.get(p["tier"], 0), p["synergyCount"]))

    char_index.append({
        "name":         char_name,
        "status":       status,
        "price":        price,
        "perkCount":    len(char_perks),
        "tierScore":    tier_score,
        "synergyScore": synergy_score,
        "totalScore":   total_score,
        "bestPerk":     best["name"],
        "bestTier":     best["tier"],
        "perks": sorted([
            {
                "id":           p["id"],
                "name":         p["name"],
                "tier":         p["tier"],
                "tierScore":    TIER_SCORE.get(p["tier"], 0),
                "synergyCount": p["synergyCount"],
                "category":     p["category"],
            }
            for p in char_perks
        ], key=lambda p: (-p["tierScore"], -p["synergyCount"])),
    })

char_index.sort(key=lambda c: -c["totalScore"])
rank = 1
for c in char_index:
    if "base game" in c["name"].lower():
        c["rank"] = 0
    else:
        c["rank"] = rank
        rank += 1

print(f"[analysis] character value index: {len(char_index)} entries")
print(f"  top 5: " + ", ".join(f"{c['name']} ({c['totalScore']})" for c in char_index[:5]))

# ── 8.5. Merge official descriptions + version history from nightlight.gg ──────
#   Source of truth for in-game wording: scripts/nightlight_perks.json
#   (produced by scripts/scrape_nightlight.py). Replaces editorial summaries
#   with the official perk description and attaches a per-perk change log.
NL_FILE = Path(__file__).parent / "perk.json"

# data.js perk name → nightlight.gg perk name (source had speech-rec / spelling errors)
NL_ALIAS = {
    "Will Make It":      "We'll make it",
    "1 2 3 4":           "ONE-TWO-THREE-FOUR!",
    "Better Boom":       "Bada Bada Boom",
    "Eyes of Belmond":   "Eyes of Belmont",
    "Soul Survivor":     "Sole Survivor",
    "Hex: Blood Favour": "Hex: Blood Favor",
    "Barbecue & Chilli": "Barbecue & Chili",
}


def _nl_norm(name):
    n = name.lower().strip()
    if ":" in n:
        n = n.split(":")[-1].strip()
    return re.sub(r"[^a-z0-9]+", "", n)


if NL_FILE.exists():
    nl_records = json.loads(NL_FILE.read_text(encoding="utf-8"))
    nl_by_full = {r["name"].lower().strip(): r for r in nl_records}
    nl_by_norm = {}
    for r in nl_records:
        nl_by_norm.setdefault(_nl_norm(r["name"]), r)

    def _nl_match(name):
        if name in NL_ALIAS:
            return nl_by_full.get(NL_ALIAS[name].lower().strip())
        return nl_by_full.get(name.lower().strip()) or nl_by_norm.get(_nl_norm(name))

    def _clean_history(vh):
        out = []
        for v in vh:
            changes = [{"type": c["type"], "text": c["text"]}
                       for c in v["changes"] if c.get("text")]
            if changes:
                out.append({"version": v["version"], "date": v["date"],
                            "changes": changes})
        return out

    merged, missing = 0, []
    for p in perks + killer_perks:
        r = _nl_match(p["name"])
        if r and r.get("description"):
            p["description"] = r["description"]
            p["versionHistory"] = _clean_history(r["version_history"])
            merged += 1
        else:
            p.setdefault("versionHistory", [])
            missing.append(p["name"])
    print(f"[nightlight] official descriptions applied: {merged} "
          f"(no source: {len(missing)} -> {missing})")
else:
    print("[nightlight] WARNING: nightlight_perks.json not found — "
          "run scripts/scrape_nightlight.py first. Descriptions unchanged.")
    for p in perks + killer_perks:
        p.setdefault("versionHistory", [])


# ── 8.6. Attach character portraits (images/CharPortraits/<Chapter>/*) ─────────
#   Killer files : K##_The<Name>_Portrait.png   Survivor: S##_<FullName>_Portrait.png
#   Match by token-set containment so "Bill Overbeck" hits "WilliamBillOverbeck",
#   "Lee Yun-Jin" hits "YunJinLee", etc. Killer names split on / and & first.
PORTRAIT_DIR = ROOT / "images" / "CharPortraits"
# data name -> in-asset spelling, only where token matching can't bridge the gap
PORTRAIT_ALIAS = {
    "Ash Williams":                 "Ashley J Williams",
    "Talita Lyra":                  "Thalita Lyra",
    "Aestri Yazar & Baermar Uraz":  "The Troupe",
    "Good Guy / Chucky":            "Yerkes",
    "Mastermind / Wesker":          "Master Mind",
    "V (Bovine Yak)":               "Vee Boonyasak",
}


def _ptokens(s):
    # split camelCase boundaries (incl. runs of capitals like "JWilliams"),
    # then on any non-alphanumeric
    s = re.sub(r"(?<=[A-Za-z])(?=[A-Z][a-z])", " ", s)   # ..JWilliams -> J Williams
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)        # Ashley|J, Leon|S
    return {t for t in re.split(r"[^A-Za-z0-9]+", s.lower()) if t}


def _attach_portraits(chars, kind):
    # index portrait files for this role: list of (tokenset, relpath)
    prefix = "K" if kind == "killer" else "S"
    # filenames: K09_<Name>_Portrait.png  or  T_UI_S51_<Name>_Portrait.png
    pat = re.compile(rf"^(?:T_UI_)?{prefix}\d+_(.*)$")
    index = []
    if PORTRAIT_DIR.exists():
        for f in PORTRAIT_DIR.rglob("*_Portrait.png"):
            stem = f.name[:-len("_Portrait.png")]
            m = pat.match(stem)
            if not m:
                continue
            newest = stem.startswith("T_UI_")             # reworked theme
            index.append((newest, _ptokens(m.group(1)),
                          f.relative_to(ROOT).as_posix()))
        index.sort(key=lambda e: not e[0])                # prefer T_UI_ on ties

    matched = 0
    for c in chars:
        name = PORTRAIT_ALIAS.get(c["name"], c["name"])
        # candidate token-sets: whole name + each / or & alternative
        cands = [_ptokens(name)]
        for part in re.split(r"[/&]", name):
            if part.strip():
                cands.append(_ptokens(part))
        hit = ""
        for _newest, toks, rel in index:
            if any(cand and cand <= toks for cand in cands):
                hit = rel
                break
        c["portrait"] = hit
        if hit:
            matched += 1
    print(f"[portraits] {kind}: {matched}/{len(chars)} matched "
          f"({len(index)} files)")
    return matched


_attach_portraits(killers, "killer")
_attach_portraits(survivors, "survivor")


# ── 9. Emit data.js ─────────────────────────────────────────────────────────────
def js_var(name, value):
    return f"const {name} = {json.dumps(value, ensure_ascii=False, indent=2)};\n"

output = (
    "// AUTO-GENERATED by scripts/build_data.py — do not edit manually\n"
    "// Source: dbd_perks_updated_F - Copy.xlsx  +  dbd_perks_updated_F.txt\n"
    "// Patch: DBD 9.6.0\n\n"
    + js_var("PERKS",        perks)
    + "\n"
    + js_var("KILLERS",      killers)
    + "\n"
    + js_var("KILLER_PERKS", killer_perks)
    + "\n"
    + js_var("SURVIVORS",    survivors)
    + "\n"
    + js_var("BUILDS",       builds)
    + "\n"
    + js_var("META",         meta)
    + "\n"
    + js_var("CHAR_VALUE",        char_index)
    + "\n"
    + js_var("KILLER_CHAR_VALUE", killer_char_index)
)

OUT.write_text(output, encoding="utf-8")
print(f"\nOK  data.js written ({OUT.stat().st_size // 1024} KB)")
print(f"  perks={len(perks)}  killers=22  survivors=52  builds=15  char_index={len(char_index)}")
