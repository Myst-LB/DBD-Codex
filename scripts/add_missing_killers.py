#!/usr/bin/env python3
"""Append 19 missing killers to Killer Tier List sheet."""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "dbd_perks_updated_F - Copy.xlsx"

NEW_KILLERS = [
    (23, 'S', 'Krasue',                    'Tethered Chains (grapple + chain hooks)',
     'Exceptional map mobility, chains through obstacles, relentless pressure',
     'New killer; nerfs expected; still being figured out by survivors',
     'Paid', '~500 AC / ~$5'),
    (24, 'A', 'Ghoul / Kaneki',            'Kagune Frenzy (multi-hit dash)',
     'Strong 1v4, Deep Wound on multiple survivors, high skill ceiling chase power',
     'Frenzy recovery leaves brief vulnerability; requires skill to maximise',
     'Paid', 'Tokyo Ghoul DLC ~$8'),
    (25, 'A', 'Singularity',               'Slipstream + Biopod teleport',
     'Unmatched map control via pods, teleport to infected survivors, great intel',
     'Moderate learning curve; flamethrower counters when survivors coordinate',
     'Paid', '~500 AC / ~$5'),
    (26, 'A', 'Dark Lord / Dracula',        'Vampiric Powers (Bat / Wolf / Armored forms)',
     'Three versatile forms, Bat for mobility, Wolf for chase, Armor for power',
     'Form switching takes time; high mental load managing all abilities',
     'Paid', 'Castlevania DLC ~$8'),
    (27, 'A', 'Artist',                    'Birds of Torment (projectile crow swarms)',
     'Anti-loop ranged pressure, punishes healing, crowd control across distance',
     'Birds can be body-blocked; power wasted at wrong loops',
     'Paid', '~500 AC / ~$5'),
    (28, 'A', 'Houndmaster',               'Scout (dog companion + fetch)',
     'Dog provides additional chase tool, searches areas independently, creates 2v1 pressure',
     'Dog pathing inconsistent; requires coordination to maximise',
     'Paid', '~500 AC / ~$5'),
    (29, 'A', 'Mastermind / Wesker',        'Virulent Bound (dash + grab/throw)',
     'Strong mobility dash, throws survivors into obstacles for bonus hits, high skill expression',
     'Dash can be dodged; must close gap before using power',
     'Paid', 'Resident Evil DLC ~$8'),
    (30, 'A', 'Lich / Vecna',              'Spellbook (Mage Hand, Fly, Dispelling Sphere)',
     'Multiple spells for varied situations; Fly grants exceptional mobility burst',
     'Spell selection is situational; spells recharge slowly',
     'Paid', 'Dungeons & Dragons DLC ~$8'),
    (31, 'B', 'Cenobite / Pinhead',        'Summons of Pain + Chain Hunt',
     'Chain Hunt escalates pressure automatically, good area denial, punishes box-hunting',
     'Survivors can solve box to pause chain hunt; requires perks to stay threatening',
     'Paid', 'Hellraiser DLC ~$8'),
    (32, 'B', 'Good Guy / Chucky',         'Scamper + Hidey-Ho (tiny stealth)',
     'Extremely hard to see, fast Scamper through obstacles, low terror radius',
     'Small hitbox can make precision difficult; survivors adapt to size quickly',
     'Paid', "Child's Play DLC ~$8"),
    (33, 'B', 'Unknown',                   'UVX (ranged projectile + teleport)',
     'Strong anti-loop ranged hit, can teleport to a UVX-struck survivor, relentless pressure',
     'UVX requires good aim; teleport telegraphed',
     'Paid', '~500 AC / ~$5'),
    (34, 'B', 'Animatronic / Springtrap',  'Dire Crank (mechanical ambush tool)',
     'Good versatility, jack-of-all-trades toolkit, low downtime between chases',
     'Average at everything; lacks explosive power of higher-tier killers',
     'Paid', 'Five Nights at Freddy\'s DLC ~$8'),
    (35, 'B', 'Dredge',                    'Nightfall + Locker teleport',
     'Locker control denies safety spots, Nightfall boosts detection and mobility',
     'Experienced survivors counter locker pathing; Nightfall inconsistent on open maps',
     'Paid', '~500 AC / ~$5'),
    (36, 'B', 'Xenomorph',                 'Crawler Mode + Tail Attack',
     'Tail attack hits through pallets, tunnel traversal, surprisingly strong at close range',
     'Flamethrower turrets hard-counter when survivors coordinate placement',
     'Paid', 'Alien DLC ~$8'),
    (37, 'B', 'Nemesis',                   'Tentacle Strike (3-tier contamination) + Zombies',
     'Zombie harassment for free, tentacle reaches through pallets, vaccine disrupts survivors',
     'Zombies inconsistent; add-ons hit-or-miss; needs 3 hits to contaminate',
     'Paid', 'Resident Evil DLC ~$8'),
    (38, 'B', 'Knight',                    'Guardia Compagnia (summon guards)',
     'Guards pressure distant gens, harassment tools, partially base-kitted improvements',
     'Guards require 50/50 wins; normal movement speed; power not immediately oppressive',
     'Paid', '~500 AC / ~$5'),
    (39, 'C', 'Onryo / Sadako',            'Condemned (TV manifestation + tape mechanic)',
     'Side-objective pressure, confuses inexperienced teams, stealth manifestation tricks',
     'Half a power at high level; stealth mediocre; counterplay is simple for good survivors',
     'Paid', 'The Ring DLC ~$8'),
    (40, 'C', 'Trickster',                 'Showstopper (knife throws + Main Event)',
     'Main Event bursts down grouped survivors, punishes body blocks, high skill ceiling',
     'No map mobility; reload from lockers; weird add-ons; known rework incoming',
     'Paid', '~500 AC / ~$5'),
    (41, 'D', 'Skull Merchant',            'Drones (scan + claw traps)',
     'Drone area control, some information gathering, occasional claw trap catches',
     'Reworked into near-useless state; half a power; F tier in Otz ranking',
     'Paid', '~500 AC / ~$5'),
]

wb = openpyxl.load_workbook(XLSX)
ws = wb['Killer Tier List']

# Find the last data row
last_row = 0
for i, row in enumerate(ws.iter_rows(), start=1):
    if row[0].value and str(row[0].value).strip().isdigit():
        last_row = i

print(f'Last data row: {last_row}')
next_row = last_row + 1

for k in NEW_KILLERS:
    ws.cell(row=next_row, column=1).value = k[0]
    ws.cell(row=next_row, column=2).value = k[1]
    ws.cell(row=next_row, column=3).value = k[2]
    ws.cell(row=next_row, column=4).value = k[3]
    ws.cell(row=next_row, column=5).value = k[4]
    ws.cell(row=next_row, column=6).value = k[5]
    ws.cell(row=next_row, column=7).value = k[6]
    ws.cell(row=next_row, column=8).value = k[7]
    print(f'  Added row {next_row}: [{k[1]}] {k[2]}')
    next_row += 1

wb.save(XLSX)
print(f'Saved: {XLSX}')
print(f'Total killers now: {last_row - 2 + len(NEW_KILLERS)}')
