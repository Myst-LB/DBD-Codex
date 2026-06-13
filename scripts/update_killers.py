#!/usr/bin/env python3
"""Rewrite all 41 killer rows in correct tier order (S > A > B > C > D)."""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "dbd_perks_updated_F - Copy.xlsx"

# Full 41-killer list ordered by Otz 9.2.0 tier placement (best → worst)
# (rank, tier, name, power, strengths, weaknesses, status, price, perk1, perk2, perk3)
ALL_KILLERS = [
    # ── S tier ──────────────────────────────────────────────────────────────
    (1,  'S', 'Nurse',
     'Blink teleport',
     'Unmatched mobility, no map safety, ends chases instantly with skill',
     'Very high skill floor; blink misses are costly',
     'Free', 'Base game',
     "A Nurse's Calling", 'Stridor', 'Thanatophobia'),
    (2,  'S', 'Blight',
     'Rush + Lethal Rush bounce',
     'Best mobility of any killer, high skill ceiling, lethal at walls and loops',
     'Steep learning curve; easy to whiff Rushes',
     'Paid', 'Archives / ~$8',
     "Dragon's Grip", 'Hex: Blood Favour', 'Hex: Undying'),
    (3,  'S', 'Krasue',
     'Tethered Chains (grapple + chain hooks)',
     'Exceptional map mobility, chains through obstacles, relentless pressure',
     'New killer; nerfs expected; survivors still learning counterplay',
     'Paid', '~500 AC / ~$5',
     'Hex: Overture of Doom', 'Ravenous', 'Wandering Eye'),

    # ── A tier ──────────────────────────────────────────────────────────────
    (4,  'A', 'Hillbilly',
     'Chainsaw sprint (insta-down)',
     'Fastest non-power traversal, shreds pallets, A+ threat in skilled hands',
     'Tight turning radius; overheat mechanic post-nerf',
     'Free', 'Base game',
     'Enduring', 'Lightborn', 'Tinkerer'),
    (5,  'A', 'Twins',
     'Charlotte + Victor split control',
     'Victor grants free pressure and map coverage, unique hook/tunnel threat',
     'Complex to master; coordination required between both bodies',
     'Paid', '~500 AC / ~$5',
     'Coup de Grâce', 'Hoarder', 'Oppression'),
    (6,  'A', 'Ghoul / Kaneki',
     'Kagune Frenzy (multi-hit dash)',
     'Strong 1v4, Deep Wound on multiple survivors, high skill ceiling chase power',
     'Frenzy recovery leaves brief vulnerability; requires skill to maximise',
     'Paid', 'Tokyo Ghoul DLC ~$8',
     'Hex: Nothing But Misery', 'Forever Entwined', 'None Are Free'),
    (7,  'A', 'Singularity',
     'Slipstream + Biopod teleport',
     'Unmatched map control via pods, teleport to infected survivors, great intel',
     'Moderate learning curve; flamethrower counters when survivors coordinate',
     'Paid', '~500 AC / ~$5',
     'Genetic Limits', 'Forced Hesitation', 'Machine Learning'),
    (8,  'A', 'Dark Lord / Dracula',
     'Vampiric Powers (Bat / Wolf / Armored forms)',
     'Three versatile forms, Bat for mobility, Wolf for chase, Armor for power',
     'Form switching takes time; high mental load managing all abilities',
     'Paid', 'Castlevania DLC ~$8',
     'Hex: Wretched Fate', 'Human Greed', 'Dominance'),
    (9,  'A', 'Spirit',
     'Phase Walk (invisible sprint)',
     'Extremely hard to read while phasing, leaves husk decoy, strong chase power',
     'Tricky to track survivors while phasing; high mental load',
     'Paid', '~500 AC / ~$5',
     'Spirit Fury', 'Hex: Haunted Ground', 'Rancor'),
    (10, 'A', 'Oni',
     'Blood Fury sprint + insta-down',
     'Best snowball killer, forgiving sprint collision, strong add-ons',
     'Slow start; needs early hits to fuel Blood Fury',
     'Paid', '~500 AC / ~$5',
     'Zanshin Tactics', 'Blood Echo', 'Nemesis'),
    (11, 'A', 'Plague',
     'Vile Purge + Corrupt Purge',
     'Efficiently injures whole team, great information tool, pressure on healers',
     'Survivors can refuse to cleanse; corrupt purge timing',
     'Paid', '~500 AC / ~$5',
     'Corrupt Intervention', 'Infectious Fright', 'Dark Devotion'),
    (12, 'A', 'Artist',
     'Birds of Torment (projectile crow swarms)',
     'Anti-loop ranged pressure, punishes healing, crowd control across distance',
     'Birds can be body-blocked; power wasted at wrong loops',
     'Paid', '~500 AC / ~$5',
     'Grim Embrace', 'Scourge Hook: Pain Resonance', 'Hex: Pentimento'),
    (13, 'A', 'Freddy / The Nightmare',
     'Dream Snares + Gen Teleport',
     'Strong gen pressure via teleport, snares slow loops, passive ability drains survivors',
     'Counter-play in waking up; somewhat passive power',
     'Paid', 'A Nightmare on Elm Street DLC ~$8',
     'Fire Up', 'Remember Me', 'Blood Warden'),
    (14, 'A', 'Houndmaster',
     'Scout (dog companion + fetch)',
     'Dog provides additional chase tool, searches areas independently, creates 2v1 pressure',
     'Dog pathing inconsistent; requires coordination to maximise',
     'Paid', '~500 AC / ~$5',
     'All-Shaking Thunder', 'Scourge Hook: Jagged Compass', 'No Quarter'),
    (15, 'A', 'Mastermind / Wesker',
     'Virulent Bound (dash + grab/throw)',
     'Strong mobility dash, throws survivors into obstacles for bonus hits, high skill expression',
     'Dash can be dodged; must close gap before using power',
     'Paid', 'Resident Evil DLC ~$8',
     'Superior Anatomy', 'Awakened Awareness', 'Terminus'),
    (16, 'A', 'Lich / Vecna',
     'Spellbook (Mage Hand, Fly, Dispelling Sphere)',
     'Multiple spells for varied situations; Fly grants exceptional mobility burst',
     'Spell selection is situational; spells recharge slowly',
     'Paid', 'Dungeons & Dragons DLC ~$8',
     'Weave Attunement', 'Languid Touch', 'Dark Arrogance'),

    # ── B tier ──────────────────────────────────────────────────────────────
    (17, 'B', 'Executioner / Pyramid Head',
     'Trails of Torment + Cages',
     'Circumvents Unbreakable/DS via cages, ranged hits through obstacles, punishes pallet drops',
     'Ranged attack cooldown; cage rescues require distance management',
     'Paid', 'Silent Hill DLC ~$8',
     'Forced Penance', 'Trail of Torment', 'Deathbound'),
    (18, 'B', 'Cenobite / Pinhead',
     'Summons of Pain + Chain Hunt',
     'Chain Hunt escalates pressure automatically, good area denial, punishes box-hunting',
     'Survivors can solve box to pause chain hunt; requires perks to stay threatening',
     'Paid', 'Hellraiser DLC ~$8',
     'Deadlock', 'Hex: Plaything', 'Scourge Hook: Gift of Pain'),
    (19, 'B', 'Good Guy / Chucky',
     'Scamper + Hidey-Ho (tiny stealth)',
     'Extremely hard to see, fast Scamper through obstacles, low terror radius',
     'Small hitbox can make precision difficult; survivors adapt to size quickly',
     'Paid', "Child's Play DLC ~$8",
     'Hex: Two Can Play', "Friends 'til the End", 'Batteries Included'),
    (20, 'B', 'Unknown',
     'UVX (ranged projectile + teleport)',
     'Strong anti-loop ranged hit, teleport to a UVX-struck survivor, relentless pressure',
     'UVX requires good aim; teleport telegraphed',
     'Paid', '~500 AC / ~$5',
     'Unbound', 'Unforeseen', 'Undone'),
    (21, 'B', 'Animatronic / Springtrap',
     'Dire Crank (mechanical ambush tool)',
     'Good versatility, jack-of-all-trades toolkit, low downtime between chases',
     'Average at everything; lacks explosive power of higher-tier killers',
     'Paid', "Five Nights at Freddy's DLC ~$8",
     'Help Wanted', 'Phantom Fear', 'Haywire'),
    (22, 'B', 'Huntress',
     'Thrown hatchets (ranged)',
     'Strong DPS with hatchets, threatening at safe loops, simple transferable skill',
     'Must reload from lockers; predictable lullaby; lost defensive-play era advantages',
     'Free', 'Base game',
     'Beast of Prey', 'Territorial Imperative', 'Hex: Huntress Lullaby'),
    (23, 'B', 'Deathslinger',
     'Spear Gun (ranged chain + Deep Wound)',
     'Strong 1v1, hits through structures, good at punishing pre-drops',
     'Limited map pressure; bugs hurt consistency; no explosive mobility',
     'Paid', '~500 AC / ~$5',
     'Gearhead', "Dead Man's Switch", 'Hex: Retribution'),
    (24, 'B', 'Hag',
     'Phantasm Traps (teleport trap)',
     'Best area control, instant teleport on trap trigger, great on indoor maps',
     'Coordinate 3-gen required; less effective vs coordinated tournament teams',
     'Free', 'Base game',
     'Hex: Devour Hope', 'Hex: Ruin', 'Hex: The Third Seal'),
    (25, 'B', 'Doctor',
     'Static Blast + Shock Therapy',
     'Great information and map pressure, madness disrupts actions, strong at loops',
     'No mobility; body blocks are significant weakness at high level',
     'Free', 'Base game',
     'Monitor & Abuse', 'Overcharge', 'Overwhelming Presence'),
    (26, 'B', 'Wraith',
     'Cloak + high uncloaked speed',
     'Very fast movement while cloaked, strong add-ons, minimal downtime between chases',
     'Must uncloak to attack; predictable at high MMR; no strong map power',
     'Free', 'Base game',
     'Bloodhound', 'Predator', 'Shadowborn'),
    (27, 'B', 'Dredge',
     'Nightfall + Locker teleport',
     'Locker control denies safety spots, Nightfall boosts detection and mobility',
     'Experienced survivors counter locker pathing; Nightfall inconsistent on open maps',
     'Paid', '~500 AC / ~$5',
     'Dissolution', 'Darkness Revealed', 'Septic Touch'),
    (28, 'B', 'Xenomorph',
     'Crawler Mode + Tail Attack',
     'Tail attack hits through pallets, tunnel traversal, surprisingly strong at close range',
     'Flamethrower turrets hard-counter when survivors coordinate placement',
     'Paid', 'Alien DLC ~$8',
     'Rapid Brutality', 'Alien Instinct', 'Ultimate Weapon'),
    (29, 'B', 'Nemesis',
     'Tentacle Strike (3-tier contamination) + Zombies',
     'Zombie harassment for free, tentacle reaches through pallets, vaccine disrupts survivors',
     'Zombies inconsistent; add-ons hit-or-miss; needs 3 hits to contaminate',
     'Paid', 'Resident Evil DLC ~$8',
     'Lethal Pursuer', 'Hysteria', 'Eruption'),
    (30, 'B', 'Knight',
     'Guardia Compagnia (summon guards)',
     'Guards pressure distant gens, harassment tools, partially base-kitted improvements',
     'Guards require 50/50 wins; normal movement speed; power not immediately oppressive',
     'Paid', '~500 AC / ~$5',
     'Nowhere to Hide', 'Hex: Face the Darkness', 'Hubris'),

    # ── C tier ──────────────────────────────────────────────────────────────
    (31, 'C', 'Onryo / Sadako',
     'Condemned (TV manifestation + tape mechanic)',
     'Side-objective pressure, confuses inexperienced teams, stealth manifestation tricks',
     'Half a power at high level; stealth mediocre; counterplay is simple for good survivors',
     'Paid', 'The Ring DLC ~$8',
     'Scourge Hook: Floods of Rage', 'Call of Brine', 'Merciless Storm'),
    (32, 'C', 'Trickster',
     'Showstopper (knife throws + Main Event)',
     'Main Event bursts down grouped survivors, punishes body blocks, high skill ceiling',
     'No map mobility; reload from lockers; weird add-ons; rework incoming',
     'Paid', '~500 AC / ~$5',
     'Starstruck', 'Hex: Crowd Control', 'No Way Out'),
    (33, 'C', 'Clown',
     'Bottles: slowdown (purple) + speed (yellow)',
     'Forces early pallet drops, strong tunneling, good loop pressure near pallets',
     'No map mobility; hook-stage changes hurt defensive playstyle',
     'Paid', '~500 AC / ~$5',
     'Bamboozle', 'Coulrophobia', 'Pop Goes the Weasel'),
    (34, 'C', 'Pig / Amanda Young',
     'Ambush + Reverse Bear Traps',
     'Built-in gen stall, traps force box searches, great noob stomper',
     'Chase power situational; trap RNG punishes poor first chases',
     'Paid', '~500 AC / ~$5',
     'Scourge Hook: Hangman\'s Trick', 'Surveillance', 'Make Your Choice'),
    (35, 'C', 'Cannibal / Bubba',
     'Extended chainsaw swing (Chainsaw Tantrum)',
     'Great noob stomper, multi-hit chainsaw, strong hook-side threat',
     'Countered by windows; add-on dependent; pallet density hurts badly',
     'Free', 'Base game',
     'Knock Out', 'Barbecue & Chilli', "Franklin's Demise"),
    (36, 'C', 'The Shape / Myers',
     'Evil Within stalking (tier system)',
     'Stealthy approach, tier-3 insta-down, pallet-break in evil within mode',
     'Stalking telegraphed; tier-3 undercooked post-rework; no mobility',
     'Paid', 'Halloween DLC ~$8',
     'Save the Best for Last', 'Play with Your Food', 'Dying Light'),
    (37, 'C', 'Legion',
     'Feral Frenzy sprint + Deep Wound',
     'Injures whole team fast, low skill floor, keeps survivors busy mending',
     'No power once survivors are injured; pallet density benefits survivors',
     'Paid', '~500 AC / ~$5',
     'Discordance', 'Mad Grit', 'Iron Maiden'),
    (38, 'C', 'Demogorgon',
     'Shred + Portals (map teleport)',
     'Balanced all-rounder, portal map control, strong Shred at close range',
     'All strengths gradually nerfed; average at everything, best at nothing',
     'Paid', 'Stranger Things DLC ~$8',
     'Surge', 'Mindbreaker', 'Cruel Limits'),

    # ── D tier ──────────────────────────────────────────────────────────────
    (39, 'D', 'Ghostface',
     'Night Shroud stalk (individual exposed)',
     'Sneaky, fast stalk, explosive burst potential when multiple survivors are 99% stalked',
     'Survivors can break stalk by looking; very limited toolkit beyond stalk burst',
     'Paid', 'Scream DLC ~$8',
     'Furtive Chase', "I'm All Ears", 'Thrilling Tremors'),
    (40, 'D', 'Trapper',
     'Bear Traps (set + trigger)',
     'Punishes unaware survivors, traps lock down areas with good map knowledge',
     'Modern pallet density destroys trap control; slowest setup; D tier in current meta',
     'Free', 'Base game',
     'Unnerving Presence', 'Brutal Strength', 'Agitation'),
    (41, 'B', 'The First / Henry Creel',
     'Test Subject #001 (Vine Attack / The Upside Down / Worldbreaker)',
     'Three interlocking abilities, Upside Down dimension for Undetectable approach, Worldbreaker mode for multi-survivor pressure',
     'Complex three-phase power; all abilities require practice to use in tandem efficiently',
     'Paid', 'Stranger Things DLC ~$8',
     'Turn Back the Clock', 'Secret Project', 'Hex: Hive Mind'),
    (42, 'D', 'Skull Merchant',
     'Drones (scan + claw traps)',
     'Drone area control, some information gathering, occasional claw trap catches',
     'Reworked into near-useless state; F tier in Otz ranking; awaiting full rework',
     'Paid', '~500 AC / ~$5',
     'THWACK!', 'Leverage', 'Game Afoot'),
]

wb = openpyxl.load_workbook(XLSX)
ws = wb['Killer Tier List']

data_row_indices = []
for i, row in enumerate(ws.iter_rows(), start=1):
    if row[0].value and str(row[0].value).strip().isdigit():
        data_row_indices.append(i)

print(f'Found {len(data_row_indices)} data rows, writing {len(ALL_KILLERS)} killers')
assert len(data_row_indices) <= len(ALL_KILLERS), \
    f"More xlsx rows ({len(data_row_indices)}) than killers ({len(ALL_KILLERS)}) — remove stale rows first"

for row_idx, k in zip(data_row_indices, ALL_KILLERS):
    ws.cell(row=row_idx, column=1).value  = k[0]   # rank
    ws.cell(row=row_idx, column=2).value  = k[1]   # tier
    ws.cell(row=row_idx, column=3).value  = k[2]   # name
    ws.cell(row=row_idx, column=4).value  = k[3]   # power
    ws.cell(row=row_idx, column=5).value  = k[4]   # strengths
    ws.cell(row=row_idx, column=6).value  = k[5]   # weaknesses
    ws.cell(row=row_idx, column=7).value  = k[6]   # status
    ws.cell(row=row_idx, column=8).value  = k[7]   # price
    ws.cell(row=row_idx, column=9).value  = k[8]   # perk 1
    ws.cell(row=row_idx, column=10).value = k[9]   # perk 2
    ws.cell(row=row_idx, column=11).value = k[10]  # perk 3
    print(f'  {k[0]:2d}. [{k[1]}] {k[2]}  |  {k[8]}, {k[9]}, {k[10]}')

# Append any new killers not covered by existing rows
if len(ALL_KILLERS) > len(data_row_indices):
    # Find the last occupied row to append after
    last_row = data_row_indices[-1] if data_row_indices else 1
    next_row = last_row + 1
    for k in ALL_KILLERS[len(data_row_indices):]:
        ws.cell(row=next_row, column=1).value  = k[0]
        ws.cell(row=next_row, column=2).value  = k[1]
        ws.cell(row=next_row, column=3).value  = k[2]
        ws.cell(row=next_row, column=4).value  = k[3]
        ws.cell(row=next_row, column=5).value  = k[4]
        ws.cell(row=next_row, column=6).value  = k[5]
        ws.cell(row=next_row, column=7).value  = k[6]
        ws.cell(row=next_row, column=8).value  = k[7]
        ws.cell(row=next_row, column=9).value  = k[8]
        ws.cell(row=next_row, column=10).value = k[9]
        ws.cell(row=next_row, column=11).value = k[10]
        print(f'  NEW {k[0]:2d}. [{k[1]}] {k[2]}  |  {k[8]}, {k[9]}, {k[10]}')
        next_row += 1

wb.save(XLSX)
print(f'\nSaved: {XLSX}')
